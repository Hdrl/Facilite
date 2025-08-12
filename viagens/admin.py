from django.contrib import admin
import zipfile
import tempfile
import os
from django.http import FileResponse
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from .models import Viagem, TransacaoFinanceira
from import_export.admin import ImportExportModelAdmin
from bs4 import BeautifulSoup as bs4
import locale
import requests, re, datetime

class UserFilteredAdmin(admin.ModelAdmin):
    """
    Admin base que filtra objetos pelo usuário logado.
    Superusuários veem tudo.
    """

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.is_superuser:
            return qs
        else:
            self.exclude = ['usuario']
        return qs.filter(usuario=request.user)

    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.usuario = request.user
        obj.save()

    def has_change_permission(self, request, obj=None):
        if obj is None or request.user.is_superuser:
            return True
        return obj.usuario == request.user

    def has_delete_permission(self, request, obj=None):
        if obj is None or request.user.is_superuser:
            return True
        return obj.usuario == request.user

def gerar_relatorio(modeladmin, request, queryset):
    if len(queryset)>1:
        return HttpResponse("Limitado a 1 relatorio por vez")
    tmp_dir = tempfile.mkdtemp()
    zip_path = os.path.join(tmp_dir, 'exportacao.zip')
    viagem = queryset[0]
    despesas = TransacaoFinanceira.objects.filter(viagem=viagem, usuario=request.user ,tipo='S').order_by('data')
    adiantamentos = TransacaoFinanceira.objects.filter(viagem=viagem, usuario=request.user, tipo='E').order_by('data')
    wb = load_workbook('relatorio/relatorio_despesas.xlsx')
    ws = wb['Despesas']
    ws['A2'] = f"Empresa: {viagem.empresa}" 
    ws['C2'] = f"Setor: {viagem.setor}"
    ws['A3'] = f"Colaborador: {viagem.colaborador}"
    ws['C3'] = f"Retorno: {viagem.retorno}"
    ws['A4'] = f"Destino: {viagem.destino}"
    ws['B4'] = f"Motivo Viagem: {viagem.motivo}"
    ws['A5'] = f"Saída: {viagem.saida}"
    #gerar cabeçalho do relatorio
    
    with zipfile.ZipFile(zip_path, 'w')as zipf:
        borda = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        fonte_hyperlink = Font(color='0563C1', underline='single', name='Arial', size=12)
        fonte_titulo = Font(bold=True, name='Arial', size=12)
        fonte_normal = Font(name='Arial', size=12)
        alinhamento = Alignment(horizontal='center', vertical='center')

        # Aplica a todas as células preenchidas
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alinhamento
                if 1 <= cell.row <= 8 and 1 <= cell.column <= 4:
                    cell.font = fonte_titulo
                else:
                    cell.font = fonte_normal
                
        for i, despesa in enumerate(despesas, start=9):
            descricao_despesa = ws.cell(row=i, column=1)
            data_despesa = ws.cell(row=i, column=2)
            nota_despesa = ws.cell(row=i, column=3)
            valor_despesa = ws.cell(row=i, column=4)
            
            if despesa.descricao:
                descricao_despesa.value = despesa.descricao.lower()
            if despesa.data:
                data_despesa.value = despesa.data.strftime('%d/%m/%Y %H:%M')
            if despesa.imagem:
                nota_despesa.value = f"=hyperlink(\"notas fiscal/{os.path.basename(despesa.imagem.path)}\", \"{os.path.basename(despesa.imagem.path)}\")"
            if despesa.valor:
                valor_despesa.value = despesa.valor
            valor_despesa.number_format = 'R$ #,##0.00'
            
            descricao_despesa.border = borda
            data_despesa.border = borda
            nota_despesa.border = borda
            valor_despesa.border = borda
            
            nota_despesa.font = fonte_hyperlink 
            
            if despesa.imagem:
                caminho = despesa.imagem.path
                nome_arquivo = os.path.basename(caminho)
                zipf.write(caminho, arcname=F"notas fiscal/{nome_arquivo}")
        
        #9 é o offset para despesas
        total_row = len(despesas) + 9
        total_cell = ws.cell(row=total_row, column=1)
        total_cell.value = "TOTAL"
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        total_cell.border = borda
        ws.cell(row=total_row, column=2).border = borda
        ws.cell(row=total_row, column=3).border = borda
        total_cell.font = fonte_titulo
        
        soma_cell = ws.cell(row=total_row, column=4)
        soma_cell.value = f"=sum(D9:D{len(despesas) + 9 - 1})"  
        soma_cell.border = borda
        
        cabecalho_row = total_row + 2
        cabecalho_adiantamento = ws.cell(row=cabecalho_row, column=1)
        cabecalho_adiantamento.value = "Descricão Do Adiantamento"
        ws.merge_cells(start_row = cabecalho_row, start_column = 1, end_row=cabecalho_row, end_column=4)
        cabecalho_adiantamento.font = fonte_titulo
        
        descricao_adiantamento = ws.cell(row=cabecalho_row + 1, column=1)
        data_adiantamento = ws.cell(row=cabecalho_row + 1, column=2)
        comprovante_adiantamento = ws.cell(row=cabecalho_row + 1, column=3)
        valor_adiantamento = ws.cell(row=cabecalho_row + 1, column=4)
        
        descricao_adiantamento.value = "Descrição"
        data_adiantamento.value = "Data"
        comprovante_adiantamento.value = "Comprovante"
        valor_adiantamento.value = "Value"
        
        cabecalho_adiantamento.border = borda
        ws.cell(row=cabecalho_row, column=2).border = borda
        ws.cell(row=cabecalho_row, column=3).border = borda
        ws.cell(row=cabecalho_row, column=4).border = borda
        descricao_adiantamento.border = borda
        data_adiantamento.border = borda
        comprovante_adiantamento.border = borda
        valor_adiantamento.border = borda
        
        cabecalho_adiantamento.font = fonte_titulo
        descricao_adiantamento.font = fonte_titulo
        data_adiantamento.font = fonte_titulo
        comprovante_adiantamento.font = fonte_titulo
        valor_adiantamento.font = fonte_titulo
        adiantamento_row = cabecalho_adiantamento.row + 1
        for i, adiantamento in enumerate(adiantamentos, cabecalho_adiantamento.row + 2):
            descricao_adiantamento = ws.cell(row=i, column=1)
            data_adiantamento = ws.cell(row=i, column=2)
            comprovante_adiantamento = ws.cell(row=i, column=3)
            valor_adiantamento = ws.cell(row=i, column=4)
            
            if adiantamento.descricao:
                descricao_adiantamento.value = adiantamento.descricao.lower()
            if adiantamento.data:
                data_adiantamento.value = adiantamento.data.strftime('%d/%m/%Y %H:%M')
            if adiantamento.imagem:
                comprovante_adiantamento.value = f"=hyperlink(\"notas fiscal/{os.path.basename(adiantamento.imagem.path)}\", \"{os.path.basename(adiantamento.imagem.path)}\")"
            if adiantamento.valor:
                valor_adiantamento.value = adiantamento.valor
            valor_adiantamento.number_format = 'R$ #,##0.00'
            
            descricao_adiantamento.border = borda
            data_adiantamento.border = borda
            comprovante_adiantamento.border = borda
            valor_adiantamento.border = borda
            
            comprovante_adiantamento.font = fonte_hyperlink
            adiantamento_row = descricao_adiantamento.row
            if adiantamento.imagem:
                caminho = adiantamento.imagem.path
                nome_arquivo = os.path.basename(caminho)
                zipf.write(caminho, arcname=F"notas fiscal/{nome_arquivo}")
            
        receber = ws.cell(row=adiantamento_row+2, column=1)
        devolver = ws.cell(row=adiantamento_row+2, column=3)
        valor_receber = ws.cell(row=adiantamento_row+3, column=1)
        valor_devolver = ws.cell(row=adiantamento_row+3, column=3)
        
        ws.merge_cells(start_row=receber.row, end_row=receber.row, start_column=1, end_column=2)
        ws.merge_cells(start_row=devolver.row, end_row=devolver.row, start_column=3, end_column=4)
        ws.merge_cells(start_row=receber.row+1, end_row=receber.row+1, start_column=1, end_column=2)
        ws.merge_cells(start_row=devolver.row+1, end_row=devolver.row+1, start_column=3, end_column=4)
        ws.cell(row=devolver.row, column=4).border = borda
        ws.cell(row=valor_devolver.row, column=4).border = borda
        ws.cell(row=receber.row, column=2).border = borda
        ws.cell(row=valor_receber.row, column=2).border = borda
        receber.border = borda
        devolver.border = borda
        valor_receber.border = borda
        valor_devolver.border = borda
           
        receber.value = "Receber"
        devolver.value = "Devolver"
        receber.font = fonte_titulo
        devolver.font = fonte_titulo
        valor_receber.value = f"=if(sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}) - D{soma_cell.row} <0, D{soma_cell.row} - sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}), 0)"
        valor_devolver.value = f"=if(sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}) - D{soma_cell.row}<0, 0, sum(D{cabecalho_adiantamento.row+2}:D{adiantamento_row}) -  D{soma_cell.row})"
        valor_devolver.number_format = 'R$ #,##0.00'
        valor_receber.number_format = 'R$ #,##0.00'
        
        gerado_em = ws.cell(row=valor_devolver.row + 2, column=1)
        ws.merge_cells(start_row=gerado_em.row, end_row=gerado_em.row, start_column=1, end_column=4)
        gerado_em.value = f"Apucarana, {datetime.datetime.now().strftime('%d de %B de %Y')} "
        
        ass_colaborador = ws.cell(row=gerado_em.row + 2, column=1)
        ass_financeiro = ws.cell(row=gerado_em.row + 2, column=3)
        valor_ass_colaborador = ws.cell(row=ass_colaborador.row+1, column=1)
        valor_ass_financeiro = ws.cell(row=ass_financeiro.row+1, column=3)
        
        ws.merge_cells(start_row=ass_colaborador.row, end_row=ass_colaborador.row, start_column=1, end_column=2)
        ws.merge_cells(start_row=ass_financeiro.row, end_row=ass_financeiro.row, start_column=3, end_column=4)
        ws.merge_cells(start_row=valor_ass_colaborador.row, end_row=valor_ass_colaborador.row, start_column=1, end_column=2)
        ws.merge_cells(start_row=valor_ass_financeiro.row, end_row=valor_ass_financeiro.row, start_column=3, end_column=4)
        
        ass_colaborador.value = "__________________________________"
        ass_financeiro.value = "__________________________________"
        valor_ass_colaborador = "Ass. Colaborador"
        valor_ass_financeiro = "Ass. Financeiro"       
        wb.save('relatorio/despesas.xlsx')
        zipf.write('relatorio/despesas.xlsx', arcname="despesas.xlsx")
    return FileResponse(open(zip_path, 'rb'), as_attachment=True, filename='despesas.zip')

def extrair_url(modeladmin, request, queryset):
    for receita in queryset:
        url = receita.nota_fiscal
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/115.0.0.0 Safari/537.36"
            )
        }
        response = requests.get(url, headers=headers)
        if not response:
            continue
        if response.status_code == 200:
            soup = bs4(response.text, "html.parser")
            #descricao
            desc_transacao = soup.find_all(id="u20")
            if desc_transacao:
                receita.descricao = desc_transacao[0].text
            #valor
            valor_transacao = soup.find_all(class_="totalNumb txtMax")
            if valor_transacao:
                receita.valor = valor_transacao[0].text.replace(',', '.')
            #data
            r_str = r"\b\d{2}/\d{2}/\d{4}\b \b\d{2}:\d{2}:\d{2}\b"
            dtime = re.search(r_str, response.text)
            if dtime:
                receita.data = datetime.datetime.strptime(dtime.group(), "%d/%m/%Y %H:%M:%S")
        receita.save()
                
@admin.register(Viagem)
class ViagemAdmin(ImportExportModelAdmin, UserFilteredAdmin):
        actions = [gerar_relatorio]

@admin.register(TransacaoFinanceira)
class TransacaoFinanceiraAdmin(ImportExportModelAdmin, UserFilteredAdmin):
        list_display = ['descricao', 'data', 'valor']
        list_filter = ['viagem']
        actions = [extrair_url]
        
        def formfield_for_foreignkey(self, db_field, request, **kwargs):
            if db_field.name == "viagem":
                if request.user.is_superuser:
                    kwargs["queryset"] = Viagem.objects.all()
                else:
                    kwargs["queryset"] = Viagem.objects.filter(usuario=request.user)
            return super().formfield_for_foreignkey(db_field, request, **kwargs)


admin.site.site_header = "Administração Viagem"
admin.site.site_title = "Administração Viagem"
admin.site.index_title = "Bem-vindo ao painel"